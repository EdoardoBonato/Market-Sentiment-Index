# -*- coding: utf-8 -*-
"""
Created on Mon Dec  4 19:47:34 2023

@author: bonated
"""

import pandas as pd
import numpy as np
import openpyxl
import plotly.express as px
import plotly.io as pio
import plotly.graph_objects as go
import os

#import merged dataset
path_merged = r"C:\your_path\merged.xlsx"
final = pd.read_excel(path_merged, sheet_name = 'Data', header = 0, na_values = "-")

#store the variables to a list
column_names = final.columns.tolist()

#subset dataset by area and sector
europe = final[(final.HQ_ctry == "USA")]
europe_cars = europe[(europe.GICS_ind_name == "Automobiles") | (europe.GICS_ind_name == "Automobile Components")]

#subsetting by year. Essentially it ends up with a collection of datasets(each one for year)
#this consider ONLY the firms for which we have values both for historical data and estimates
data = {}
for year in final["FY"].unique():
    df_name = "final_cars_" + str(year)
    t = europe_cars[europe_cars.FY == year]
    data[df_name] = t
common_names = data["final_cars_2023"]["company_name"].unique()
common_names = common_names.tolist() 

dataframes_dic = {}
for key, dataframe in data.items():
        ebit_name = "EBIT_" + key[11:15]
        ebit_std_name = "EBIT_std_" + key[11:15]
        meancost = dataframe.pivot(index= 'company_code', columns='FY', values='EBIT')
        stdcost = dataframe.pivot(index = 'company_code', columns='FY', values='EBITStdDev')
        dataframes_dic[ebit_name] = meancost
        dataframes_dic[ebit_std_name] = stdcost
    
dataframes_dic.pop('EBIT_2017')      

#calculate the weights
revenue_sector = data['final_cars_2018']['revenue'].sum()
weights = {}
for index, row in data['final_cars_2018'].iterrows():
    weights[row['company_code']] = (row['revenue'] / revenue_sector)
    
weights = pd.DataFrame(list(weights.items()), columns = ['company_code', 'weights'])   

#construct the final dataset for the index, it contains cost and  cost std for each year and each company
company_names = data["final_cars_2023"].pivot(index = 'company_code', columns = 'FY', values = 'company_name')
EBIT_df = pd.DataFrame(data= None, index = None)    
EBIT_df["company_name"] = company_names
for key, dataframe in dataframes_dic.items():
    EBIT_df[key] = dataframe
    
#drop the useless columns
EBIT_df = EBIT_df.drop(columns =["EBIT_std_2017", "EBIT_std_2018", "EBIT_std_2019", "EBIT_std_2020",
                             "EBIT_std_2021", "EBIT_std_2022"])
    
#assign index mean titles
index_mean_titles = [2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025]
 
rfrnc_yr = EBIT_df['EBIT_2018']
for year in index_mean_titles:
        indx_nm = 'index_' + str(year)
        index_std_nm = 'index_std_' + str(year)
        cst_yr = 'EBIT_' + str(year)
        # Determine the sign of the original values
        EBIT_df[indx_nm] = EBIT_df.apply(lambda row : (row[cst_yr] / abs(row['EBIT_2018']))*100 if row['EBIT_2018'] > 0 else 
                                         (row[cst_yr] / abs(row['EBIT_2018'])*100) + 200, axis = 1 )
        if int(year) >= 2023:
            EBIT_std = 'EBIT_std_' + str(year)
            EBIT_df[index_std_nm] = (EBIT_df[EBIT_std]/EBIT_df[cst_yr])*100

#insert the weights
EBIT_df = pd.merge(EBIT_df, weights, on = 'company_code', how = 'left')

#standard deviation 
index_std = 'index_std'
for column in EBIT_df.columns:
        if index_std in column:
            EBIT_df[column] = EBIT_df[column].apply(lambda x : EBIT_df.loc[EBIT_df[column] != x, column].mean() if x == 0 else x)

#this option is to have balanced panel
#drop the companies which have not all values
#take account of the companies in a list : excluded_companies
EBIT_df = EBIT_df.dropna()
#dealing with outliers
for year in index_mean_titles:
        name = 'index_' + str(year)
        EBIT_df[name] = np.clip(EBIT_df[name], EBIT_df[name].quantile(0.10), EBIT_df[name].quantile(0.90, interpolation='lower'))

#calculate index mean and std
index_mean = []
for column in EBIT_df.columns[-13:-1]:
    if "std" not in column:
        index_mean.append(EBIT_df[column].mean(skipna = True))
    
index_std = [0, 0, 0, 0, 0]
for column in EBIT_df.columns[-6 :-1]:
        if "std" in column:
            index_std.append(EBIT_df[column].mean(skipna = True))
    
   
#weighted version
index_mean_weighted = []        
for column in EBIT_df.columns[-13:-1]:
        if "std" not in column:
            index_mean_weighted.append(np.average(EBIT_df[column], weights = EBIT_df['weights']))

index_std_weighted = [0, 0, 0, 0, 0]
for column in EBIT_df.columns[-6:-1]:
    if "std" in column:
        index_std_weighted.append(np.average(EBIT_df[column], weights = EBIT_df['weights']))
        
        
#export all to excel
file_path_1 = r'C:\Users\edobo\OneDrive\Desktop\THINGS TO SEND\EBIT_balanced_index_US_with_outliers.xlsx'
writer = pd.ExcelWriter(file_path_1, engine="xlsxwriter", mode='w')
EBIT_df.to_excel(writer, sheet_name = 'data')
writer.close()
workbook = openpyxl.load_workbook(file_path_1)
sheet = workbook.create_sheet("index")
for col_num, value in enumerate(index_mean_titles, start=1):
     sheet.cell(row = 1, column= 1 + col_num, value=value)
for col_num, value in enumerate(index_mean, start=1):
     sheet.cell(row = 2, column= 1 + col_num, value=value)
for col_num, value in enumerate(index_std, start=1):
     sheet.cell(row = 3, column= 1 + col_num, value=value)
for col_num, value in enumerate(index_mean_titles, start=1):
     sheet.cell(row = 5, column= 1 + col_num, value=value)
for col_num, value in enumerate(index_mean_weighted, start=1):
     sheet.cell(row = 6, column= 1 + col_num, value=value)
for col_num, value in enumerate(index_std_weighted, start=1):
     sheet.cell(row = 7, column= 1 + col_num, value=value)
 
sheet.cell(row = 1, column = 1, value = "year")
sheet.cell(row= 2, column = 1, value = "index_mean")
sheet.cell(row = 3, column = 1, value = "index_std")

sheet.cell(row = 5, column = 1, value = "year_weighted")
sheet.cell(row= 6, column = 1, value = "index_mean_weighted")
sheet.cell(row = 7, column = 1, value = "index_std_weighted")
 
workbook.save(file_path_1)
workbook.close()

#graph construction
x = index_mean_titles
x_rev = x[::-1]

y_ub = [index_mean_weighted[value] + index_std_weighted[value] for value in range(0, len(index_mean_weighted))]
y_lb = [index_mean_weighted[value] - index_std_weighted[value] for value in range(0, len(index_mean_weighted))]
y_lb = y_lb[::-1]

transition_point = 5
fig1 = go.Figure()
fig1.add_trace(go.Scatter(
    x = x[transition_point-1:] + x_rev[:transition_point-1],
    y = y_ub[transition_point-1:] + y_lb[:transition_point-1],
    fill = 'toself',fillcolor = 'rgba(0,100,80,0.2)',
    line_color ='rgba(255,255,255,0)', 
    showlegend=False
    ))
fig1.add_trace(go.Scatter(
   x = x[:transition_point] + x_rev[:transition_point],
   y = index_mean_weighted[:transition_point],
   line_color='rgb(0,100,80)',
   name = 'profitability_actual'
   ))
fig1.add_trace(go.Scatter(
    x = x[transition_point-1:] + x_rev[transition_point-1:],
    y = index_mean_weighted[transition_point-1:],
    line=dict(color='rgb(0,100,80)', dash='dot'),
    name = 'profitability_forecasts'
   ))
 
fig1.update_xaxes(title_text = "year")
fig1.update_yaxes(title_text = "Index (2018=100, nr of companies ")
fig1.update_layout(title_text = "Automotive: profitability_us,  balanced panel, weighted")
fig1.update_layout(yaxis = dict( range=[0, 150], dtick = 20))
fig1.write_html(r'C:\your_path\profitability_weighted_USA.html')



