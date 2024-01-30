# -*- coding: utf-8 -*-
"""
Created on Tue Nov 28 16:06:29 2023

@author: bonated
"""

#import useful packages
import pandas as pd
import numpy as np
import openpyxl
import plotly.express as px
import plotly.io as pio
import plotly.graph_objects as go
import os
#---personal modules
import sys
sys.path.append(r"C:\Users\edobo\OneDrive\Desktop\THINGS TO SEND")
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#INDEX FOR COSTS
#reindexing the Dataframe, select only the variable Total Cost and Revenue Standard Dviation. Output : DataFrames

def index_creation(data = None,balanced = True, variable = None, area = None):
    
    #import module and create variables
    from correspondence_codes import correspondence_code
    dataframes_dic = None
    index_mean = None
    index_std = None
    excluded_companies = None
    cost_df = None
    rev_df = None
    data_geo = None
    #define the valid answers
    valid_answers = ['cost', 'revenue']
    if variable not in valid_answers:
       raise ValueError("Invalid value for 'variable'. Please choose from: 'cost', 'revenue'")
    
    if area == 'USA' or area == 'United States of America' or area == 'US' or area == 'us' :
        area = 'USA'
        data_geo = data[(data.HQ_ctry == area )]
    if area == 'EU' or area == 'European Union' or area == 'Europe' or area == 'eu':
        area = 'EU'
        data_geo = data[data.europe == 1]
    if area == None:
        raise ValueError('Please insert a correct area')
    #select the sector
    data_sector = data_geo[(data_geo.GICS_ind_name == "Automobiles") | (data_geo.GICS_ind_name == "Automobile Components")]
    #subsetting by year. It ends up with a collection of datasets(each one for year)
    #this consider ONLY the firms for which we have values both for historical data and estimates
    data_final = {}
    for year in data["FY"].unique():
        df_name = "final_cars_" + str(year)
        t = data_sector[data_sector.FY == year]
        data_final[df_name] = t
    common_names = data_final["final_cars_2023"]["company_name"].unique()
    common_names = common_names.tolist()
    #calculate the weights
    revenue_sector = data_final['final_cars_2018']['revenue'].sum()
    weights = {}
    for index, row in data_final['final_cars_2018'].iterrows():
        weights[row['company_code']] = (row['revenue'] / revenue_sector)
    weights = pd.DataFrame(list(weights.items()), columns = ['company_code', 'weights'])    
    

    #variable cost chosen
    if variable == 'cost':
        dataframes_dic = {}
        for key, dataframe in data_final.items():
            cost_name = "cost_" + key[11:15]
            cost_std_name = "cost_std_" + key[11:15]
            meancost = dataframe.pivot(index= 'company_code', columns='FY', values='cost')
            stdcost = dataframe.pivot(index = 'company_code', columns='FY', values='COGS_std')
            dataframes_dic[cost_name] = meancost
            dataframes_dic[cost_std_name] = stdcost
        
        dataframes_dic.pop('cost_2017')      
        
        #construct the final dataset for the index, it contains cost and  cost std for each year and each company
        company_names = data_final["final_cars_2023"].pivot(index = 'company_code', columns = 'FY', values = 'company_name')
        cost_df = pd.DataFrame(data= None, index = None)    
        cost_df["company_name"] = company_names
        
        for key, dataframe in dataframes_dic.items():
            cost_df[key] = dataframe
        
        #drop the useless columns
        cost_df = cost_df.drop(columns =["cost_std_2017", "cost_std_2018", "cost_std_2019", "cost_std_2020",
                                 "cost_std_2021", "cost_std_2022"])
        
    
        #assign index mean titles
        index_mean_titles = [2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025]
        
        #index creation COST
        rfrnc_yr = cost_df['cost_2018']
        for year in index_mean_titles:
            indx_nm = 'index_' + str(year)
            index_std_nm = 'index_std_' + str(year)
            cst_yr = 'cost_' + str(year)
            cost_df[indx_nm] = (cost_df[cst_yr]/rfrnc_yr)*100
            if int(year) >= 2023:
                cst_std = 'cost_std_' + str(year)
                cost_df[index_std_nm] = (cost_df[cst_std]/cost_df[cst_yr])*100
      
        #impute standard deviation when equals 0 
        index_std = 'index_std'
        for column in cost_df.columns:
            if index_std in column:
                cost_df[column] = cost_df[column].apply(lambda x : cost_df.loc[cost_df[column] != x, column].mean() if x == 0 else x)
        
        cost_df = pd.merge(cost_df, weights, on = 'company_code', how = 'left')

        #this option is to have balanced panel
        #drop the companies which have not all values
        #take account of the companies in a list : excluded_companies
        if balanced == True:
            excluded_companies = []
            for index, row in cost_df.iterrows():
                if row.isna().any() == True:
                    excluded_companies.append(index)
                    excluded_companies.append(row.isna().sum())
                    excluded_companies_df = pd.DataFrame(excluded_companies)
            cost_df = cost_df.dropna()
        
        #dealing with outliers
        for year in index_mean_titles:
            name = 'index_' + str(year)
            cost_df[name] = np.clip(cost_df[name], cost_df[name].quantile(0.10), cost_df[name].quantile(0.90, interpolation='lower'))
         
        #calculate index mean and std
        index_mean = []
        for column in cost_df.columns[-13:-1]:
            if "std" not in column:
                index_mean.append(cost_df[column].mean(skipna = True))
        
        index_std = [0, 0, 0, 0, 0]
        for column in cost_df.columns[-6 :-1]:
            if "std" in column:
                index_std.append(cost_df[column].mean(skipna = True))
        
        #weighted version
        index_mean_weighted = []        
        for column in cost_df.columns[-13:-1]:
            if "std" not in column:
                index_mean_weighted.append(np.average(cost_df[column], weights = cost_df['weights']))

        index_std_weighted = [0, 0, 0, 0, 0]
        for column in cost_df.columns[-6:-1]:
            if "std" in column:
                index_std_weighted.append(np.average(cost_df[column], weights = cost_df['weights']))


    #variable revenue chosen
    if variable == 'revenue':
        dataframes_dic = {}
        for key, dataframe in data_final.items():
            rev_name = "revenue_" + key[11:15]
            rev_std_name = "revenue_std_" + key[11:15]
            meanrev = dataframe.pivot(index= 'company_code', columns='FY', values='revenue')
            stdrev = dataframe.pivot(index = 'company_code', columns='FY', values='revenue_std')
            dataframes_dic[rev_name] = meanrev
            dataframes_dic[rev_std_name] = stdrev
        dataframes_dic.pop('revenue_2017')  
        # construct the final dataset for the index, it contains Total Revenue and Revenue Stadnard Deviation for each year and each company
        company_names = data_final["final_cars_2023"].pivot(index = 'company_code', columns = 'FY', values = 'company_name')
        print(company_names)
        rev_df = pd.DataFrame(data= None, index = None)    
        rev_df["company_name"] = company_names
        for key, dataframe in dataframes_dic.items():
            rev_df[key] = dataframe
        #drop the useless columns
        rev_df = rev_df.drop(columns =["revenue_std_2017", "revenue_std_2018", "revenue_std_2019", "revenue_std_2020",
                                       "revenue_std_2021", "revenue_std_2022"])
        
        
        #assign index mean titles
        index_mean_titles = [2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025]
        
        #index mean creation
        rfrnc_yr = rev_df['revenue_2018']
        for year in index_mean_titles:
            indx_nm = 'index_' + str(year)
            index_std_nm = 'index_std_' + str(year)
            rev_yr = 'revenue_' + str(year)
            rev_df[indx_nm] = (rev_df[rev_yr]/rfrnc_yr)*100
            if int(year) >= 2023:
                rev_std = 'revenue_std_' + str(year)
                rev_df[index_std_nm] = (rev_df[rev_std]/rev_df[rev_yr])*100
        
        #standard deviation
        index_std = 'index_std'
        for column in rev_df.columns:
           if index_std in column:
               rev_df[column] = rev_df[column].apply(lambda x : rev_df.loc[rev_df[column] != x, column].mean() if x == 0 else x)
       
        rev_df = pd.merge(rev_df, weights, on = 'company_code', how = 'left')
        #this option is to have balanced panel
        #drop the companies which have not all values
        #take account of the companies in a list : excluded_companies
        if balanced == True:
            excluded_companies = []
            for index, row in rev_df.iterrows():
                if row.isna().any() == True:
                    excluded_companies.append(index)
                    excluded_companies.append(row.isna().sum())
                    excluded_companies_df = pd.DataFrame(excluded_companies)
            rev_df = rev_df.dropna()

        #dealing with outliers
        for year in index_mean_titles:
            name = 'index_' + str(year)
            rev_df[name] = np.clip(rev_df[name], rev_df[name].quantile(0.10), rev_df[name].quantile(0.90, interpolation='lower'))

        #index construction
        index_mean = []
        for column in rev_df.columns[-13:-1]:
            if "std" not in column:
                index_mean.append(rev_df[column].mean(skipna = True))
        index_std = [0, 0, 0, 0, 0]
        for column in rev_df.columns[-6 :-1]:
            if "std" in column:
                index_std.append(rev_df[column].mean(skipna = True))
                
        #weighted
        index_mean_weighted = []        
        for column in rev_df.columns[-13:-1]:
            if "std" not in column:
                index_mean_weighted.append(np.average(rev_df[column], weights = rev_df['weights']))

        index_std_weighted = [0, 0, 0, 0, 0]
        for column in rev_df.columns[-6:-1]:
            if "std" in column:
                index_std_weighted.append(np.average(rev_df[column], weights = rev_df['weights']))
        
    #export to _excel 
    if balanced == True:
       file_path_1 = r'C:\Users\edobo\OneDrive\Desktop\THINGS TO SEND' + '\\' + variable + '_' + area + '_balanced_index.xlsx'
    if balanced == False:
        file_path_1 = r'C:\Users\edobo\OneDrive\Desktop\THINGS TO SEND' + '\\' + variable + '_' + area + '_unbalanced_index.xlsx'
    writer = pd.ExcelWriter(file_path_1, engine="xlsxwriter", mode='w')
    if variable == 'cost':
        cost_df.to_excel(writer, sheet_name = 'data')
    if variable == 'revenue':
        rev_df.to_excel(writer, sheet_name = 'data')
    writer.close()
    '''    
    #this give the NACE names and codes which corresponds to the GICS name used 
    # correspondence between GICS and NACE
    correspondence = correspondence_code('GICS_ind_name', 'NACE_name')
    correspondence_code = []
    for gics in data["GICS_ind_name"].unique():
          for GICS, NACE in correspondence.items():
              if gics == GICS:
                  t = (GICS, NACE)
                  correspondence_code.append(t)
    correspondence_code_df = pd.DataFrame(correspondence_code)
    #excluded_companies_df.to_excel(writer, sheet_name = 'excluded_companies')
    correspondence_code_df.to_excel(writer, sheet_name = "correspondence")
    writer.close()
    '''
    workbook = openpyxl.load_workbook(file_path_1)
    sheet = workbook.create_sheet("index")
    for col_num, value in enumerate(index_mean_titles, start=1):
        sheet.cell(row = 1, column= 1 + col_num, value=value)
    for col_num, value in enumerate(index_mean, start=1):
        sheet.cell(row = 2, column= 1 + col_num, value=value)
    for col_num, value in enumerate(index_std, start=1):
        sheet.cell(row = 3, column= 1 + col_num, value=value)
    for col_num, value in enumerate(index_mean_titles, start=1):
        sheet.cell(row = 5, column= 1 + col_num, value=value)
    for col_num, value in enumerate(index_mean_weighted, start=1):
        sheet.cell(row = 6, column= 1 + col_num, value=value)
    for col_num, value in enumerate(index_std_weighted, start=1):
        sheet.cell(row = 7, column= 1 + col_num, value=value)
    
    sheet.cell(row = 1, column = 1, value = "year")
    sheet.cell(row= 2, column = 1, value = "index_mean")
    sheet.cell(row = 3, column = 1, value = "index_std")
    
    sheet.cell(row = 5, column = 1, value = "year_weighted")
    sheet.cell(row= 6, column = 1, value = "index_mean_weighted")
    sheet.cell(row = 7, column = 1, value = "index_std_weighted")
    
    workbook.save(file_path_1)
    workbook.close()
